import unittest
from io import BytesIO

from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from sqlalchemy.pool import StaticPool

from app.api.export import router
from app.database import Base, get_db
from app.models import Patent, PatentDatabase
from app.services.export_service import ExportService
from app.services.view_service import ViewService


class ExportApiTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.db = Session(self.engine)
        database = PatentDatabase(name="导出测试库", code="EXPORT_TEST")
        self.db.add(database)
        self.db.flush()
        self.db.add_all([
            Patent(title="授权专利", legal_status="granted", database_id=database.id),
            Patent(title="待审专利", legal_status="pending", database_id=database.id),
        ])
        self.db.commit()
        self.database_id = database.id

        app = FastAPI()
        app.include_router(router)
        app.dependency_overrides[get_db] = lambda: self.db
        self.client = TestClient(app)

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_csv_exports_selected_filtered_fields_with_bom(self):
        response = self.client.post("/export/csv", json={
            "database_id": self.database_id,
            "field_keys": ["title", "legal_status"],
            "filters": {"legal_status": {"eq": "granted"}},
        })
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"\xef\xbb\xbf"))
        content = response.content.decode("utf-8-sig")
        self.assertIn("标题,法律状态", content)
        self.assertIn("授权专利,授权", content)
        self.assertNotIn("待审专利", content)

    def test_excel_exports_selected_fields_and_freezes_header(self):
        response = self.client.post("/export/excel", json={
            "database_id": self.database_id,
            "field_keys": ["title"],
        })
        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook["专利数据"]
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(sheet["A1"].value, "标题")
        self.assertEqual(sheet.max_row, 3)

    def test_system_work_file_templates_are_listed_and_used_by_export(self):
        ViewService.ensure_default_business_views(self.db, self.database_id)
        templates = ExportService.ensure_default_templates(self.db, self.database_id)
        self.assertEqual(
            {template.template_key for template in templates},
            {
                "risk_meeting_excel",
                "ip_application_control_excel",
                "patent_analysis_work_file",
                "daily_patent_accumulation_csv",
            },
        )

        listed = self.client.get("/export/templates", params={"database_id": self.database_id})
        self.assertEqual(listed.status_code, 200)
        self.assertEqual(len(listed.json()), 4)

        csv_template = next(item for item in listed.json() if item["output_format"] == "csv")
        response = self.client.post("/export/csv", json={
            "database_id": self.database_id,
            "template_id": csv_template["id"],
        })
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b"\xef\xbb\xbf"))
        self.assertIn("标题", response.content.decode("utf-8-sig"))


if __name__ == "__main__":
    unittest.main()
